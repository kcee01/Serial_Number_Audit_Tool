import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import threading
import queue

# === Constants ===
LOG_DIR = os.path.join(os.environ["USERPROFILE"], "Downloads", "sample", "logs")
os.makedirs(LOG_DIR, exist_ok=True)
OpenFolder_DIR = r"C:\Users\ckeabetswe\Downloads\sample"
os.makedirs(LOG_DIR, exist_ok=True)
print("🔍 Absolute LOG_DIR:", LOG_DIR)


# === Logging ===
def log_event(message, log_queue):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_message = f"[{timestamp}] {message}"
    print(log_message)
    log_queue.put(log_message)


# === Excel Helpers ===
def find_column_index(ws: Worksheet, header: str) -> int:
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value and str(col[0].value).strip().lower() == header.lower():
            return col[0].column
    raise ValueError(f"Column '{header}' not found in sheet '{ws.title}'")


def update_progress(step, total, progress_var):
    progress_var.set(int((step / total) * 100))


def validate_excel_columns(df_prev, df_curr):
    for col in ["Serial Number"]:
        if col not in df_prev.columns or col not in df_curr.columns:
            raise KeyError(f"Missing '{col}' in input files")


# === Main Logic ===
def sync_bw_start_meter(ws_curr, ws_prev, log_list, step, total, log_queue):
    log_event("🔄 Syncing B/W Start Meter...", log_queue)
    bw_col = find_column_index(ws_curr, "B/W Start Meter")
    end_col = find_column_index(ws_prev, "End Meter")

    for row in range(2, ws_curr.max_row + 1):
        value = ws_prev.cell(row=row, column=end_col).value
        cell = ws_curr.cell(row=row, column=bw_col)
        cell.value = value
        log_list.append({
            "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Cell": cell.coordinate,
            "Value": value
        })

    log_path = os.path.join(LOG_DIR, f"bw_start_meter_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    pd.DataFrame(log_list).to_excel(log_path, index=False)
    log_event(f"📄 B/W Start Meter log saved to {log_path}", log_queue)
    log_event("✅ Synced B/W Start Meter.", log_queue)
    return step + 1


def match_serial_numbers(ws, df_curr, df_prev, step, total, progress_var, log_queue):
    log_event("🔍 Matching Serial Numbers...", log_queue)
    serial_set = set(df_prev["Serial Number"].dropna().astype(str).str.replace(" ", "").str.strip())
    serial_idx = find_column_index(ws, "Serial Number")
    insert_idx = serial_idx + 1
    ws.insert_cols(insert_idx)
    ws.cell(row=1, column=insert_idx, value="Nashua Serial Number")

    matched, matched_serials, unmatched_serials = 0, [], []

    for row in range(2, ws.max_row + 1):
        val_raw = ws.cell(row=row, column=serial_idx).value
        if not val_raw:
            continue
        val = ''.join(str(val_raw).split())
        if val in serial_set:
            ws.cell(row=row, column=insert_idx).value = val
            matched_serials.append(val)
            matched += 1
        else:
            ws.cell(row=row, column=serial_idx).value = ""
            unmatched_serials.append(val)

    log_event(f"✅ Matching complete. Matched: {matched}, Unmatched: {len(unmatched_serials)}", log_queue)
    return step + 1, matched_serials, unmatched_serials


def handle_duplicates(ws, log_queue):
    log_event("🔎 Checking for duplicates...", log_queue)
    serial_col_idx = find_column_index(ws, "Serial Number")
    headers = [cell.value for cell in ws[1]]
    seen, duplicates = set(), []

    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=serial_col_idx).value
        if val in seen:
            duplicates.append([ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)])
            ws.cell(row=row, column=serial_col_idx).value = ""
        else:
            seen.add(val)

    if duplicates:
        dup_path = os.path.join(LOG_DIR, f"duplicates_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        pd.DataFrame(duplicates, columns=headers).to_excel(dup_path, index=False)
        log_event(f"⚠️ Duplicates saved to {dup_path}", log_queue)

    log_event(f"🧼 Blank-celled {len(duplicates)} duplicate entries.", log_queue)


def delete_column(ws, col_name, log_queue):
    log_event(f"🧹 Deleting column '{col_name}'...", log_queue)
    for col in ws.iter_cols(1, ws.max_column):
        if col[0].value == col_name:
            ws.delete_cols(col[0].column)
            log_event(f"🗑️ Deleted column '{col_name}'.", log_queue)
            return True
    log_event(f"⚠️ Column '{col_name}' not found.", log_queue)
    return False


def run_audit(prev_path, curr_path, progress_var, log_queue):
    try:
        log_event("🚀 Starting audit...", log_queue)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        df_prev = pd.read_excel(prev_path)
        df_curr = pd.read_excel(curr_path, sheet_name="DATA")
        df_prev.columns = df_prev.columns.str.strip()
        df_curr.columns = df_curr.columns.str.strip()

        validate_excel_columns(df_prev, df_curr)

        wb_curr = load_workbook(curr_path)
        ws_curr = wb_curr["DATA"]
        ws_prev = load_workbook(prev_path).active

        handle_duplicates(ws_curr, log_queue)

        total_steps = 5
        step = 0

        bw_log = []
        step = sync_bw_start_meter(ws_curr, ws_prev, bw_log, step, total_steps, log_queue)
        update_progress(step, total_steps, progress_var)

        step, matched, unmatched = match_serial_numbers(ws_curr, df_curr, df_prev, step, total_steps, progress_var, log_queue)
        update_progress(step, total_steps, progress_var)

        delete_column(ws_curr, "Nashua Serial Number", log_queue)
        step += 1
        update_progress(step, total_steps, progress_var)

        matched_path = os.path.join(LOG_DIR, f"matched_serials_{timestamp}.csv")
        unmatched_path = os.path.join(LOG_DIR, f"unmatched_serials_{timestamp}.csv")
        pd.Series(matched, name="Matched Serials").to_csv(matched_path, index=False)
        pd.Series(unmatched, name="Unmatched Serials").to_csv(unmatched_path, index=False)
        log_event(f"📄 Matched & unmatched serials exported.", log_queue)

        wb_curr.save(curr_path)

        log_event("✅ Audit completed successfully.", log_queue)
        status_label.config(text="Audit completed.", fg="green")
        open_logs_btn.config(state="normal")
        update_progress(total_steps, total_steps, progress_var)

    except Exception as e:
        log_event(f"❌ Audit failed: {e}", log_queue)
        messagebox.showerror("Error", str(e))
        status_label.config(text="Audit failed.", fg="red")
        open_logs_btn.config(state="normal")


def start_audit():
    prev_path = prev_file_var.get()
    curr_path = curr_file_var.get()
    if not prev_path or not curr_path:
        messagebox.showerror("Missing File", "Please select both files.")
        return

    log_queue = queue.Queue()
    threading.Thread(target=run_audit, args=(prev_path, curr_path, progress_var, log_queue), daemon=True).start()

    def update_console():
        while not log_queue.empty():
            console.insert(tk.END, log_queue.get_nowait() + "\n")
            console.see(tk.END)
        root.after(100, update_console)

    update_console()


# === GUI ===
root = tk.Tk()
root.title("📊 Serial Number Audit Tool")
root.geometry("800x600")
root.columnconfigure(1, weight=1)
root.rowconfigure(5, weight=1)

prev_file_var = tk.StringVar()
curr_file_var = tk.StringVar()
progress_var = tk.IntVar()

tk.Label(root, text="Previous File:").grid(row=0, column=0, sticky="e")
tk.Entry(root, textvariable=prev_file_var).grid(row=0, column=1, sticky="ew")
tk.Button(root, text="Browse", command=lambda: prev_file_var.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]))).grid(row=0, column=2)

tk.Label(root, text="Current File:").grid(row=1, column=0, sticky="e")
tk.Entry(root, textvariable=curr_file_var).grid(row=1, column=1, sticky="ew")
tk.Button(root, text="Browse", command=lambda: curr_file_var.set(filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")]))).grid(row=1, column=2)

tk.Button(root, text="Run Audit", command=start_audit, bg="blue", fg="white").grid(row=2, column=1)
status_label = tk.Label(root, text="", fg="green")
status_label.grid(row=3, column=0, columnspan=2, sticky="w")

open_logs_btn = tk.Button(root, text="📁 Open Logs Folder", command=lambda: os.startfile(OpenFolder_DIR ))
open_logs_btn.grid(row=3, column=2, sticky="e")
open_logs_btn.config(state="disabled")

progress_bar = ttk.Progressbar(root, variable=progress_var, maximum=100)
progress_bar.grid(row=4, column=0, columnspan=3, sticky="ew")

console_frame = tk.Frame(root)
console_frame.grid(row=5, column=0, columnspan=3, sticky="nsew")
console = tk.Text(console_frame, bg="black", fg="lime", font=("Courier", 10), wrap="none")
console.grid(row=0, column=0, sticky="nsew")

scroll_y = tk.Scrollbar(console_frame, orient="vertical", command=console.yview)
scroll_x = tk.Scrollbar(console_frame, orient="horizontal", command=console.xview)
console.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
scroll_y.grid(row=0, column=1, sticky="ns")
scroll_x.grid(row=1, column=0, sticky="ew")

console_frame.columnconfigure(0, weight=1)
console_frame.rowconfigure(0, weight=1)

root.mainloop()
